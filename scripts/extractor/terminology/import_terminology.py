from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from scripts.extractor.pipeline.terminology_normalizer import normalize_term_key


ASKMED_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_OUTPUT = ASKMED_ROOT / "data" / "terminology" / "terminology.sqlite"


def resolve_from_root(path: Path) -> Path:
    return path if path.is_absolute() else ASKMED_ROOT / path


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def split_synonyms(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    for delimiter in ("|", "；", ";", "、"):
        text = text.replace(delimiter, "|")
    return [item.strip() for item in text.split("|") if item.strip()]


def is_usable_display_name(value: str) -> bool:
    text = str(value or "").strip()
    if not text or len(text) > 30:
        return False
    if any(marker in text for marker in (".", "^", "{", "}")):
        return False
    return bool(re.search(r"[\u4e00-\u9fff]", text))


def choose_better_preferred(current: str | None, candidate: str) -> str | None:
    if not is_usable_display_name(candidate):
        return current
    if not current or not is_usable_display_name(current) or len(candidate) < len(current):
        return candidate
    return current


def read_table(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        delimiter = "\t" if suffix in {".tsv", ".txt"} else None
        return read_delimited(path, delimiter)
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SystemExit("Reading XLSX requires openpyxl. Install it or provide CSV/TSV.") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]
    raise ValueError(f"Unsupported terminology file type: {path}")


def read_delimited(path: Path, delimiter: str | None) -> list[dict[str, Any]]:
    encodings = ("utf-8-sig", "utf-8", "gb18030")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                dialect = csv.Sniffer().sniff(sample) if delimiter is None and sample else csv.excel
                if delimiter is not None:
                    dialect.delimiter = delimiter
                return list(csv.DictReader(f, dialect=dialect))
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    if last_error:
        raise last_error
    return []


def pick(row: dict[str, Any], candidates: tuple[str, ...]) -> str:
    lowered = {str(key).strip().lower(): key for key in row}
    for candidate in candidates:
        key = lowered.get(candidate.lower())
        if key is not None:
            value = row.get(key)
            if value is not None and str(value).strip():
                return str(value).strip()
    return ""


def iter_icd_terms(path: Path) -> Iterable[dict[str, Any]]:
    for row in read_table(path):
        code = pick(row, ("code", "icd_code", "diagnosis_code", "疾病诊断代码", "编码", "代码"))
        name = pick(row, ("preferred_name", "name", "disease", "diagnosis_name", "疾病诊断名称", "疾病名称", "名称", "诊断名称"))
        synonyms = pick(row, ("synonyms", "alias", "aliases", "同义词", "别名"))
        if not code or not name:
            continue
        yield {
            "terminology": "ICD-10-CN",
            "semantic_type": "disease",
            "code": code,
            "preferred_name": name,
            "aliases": [name, *split_synonyms(synonyms)],
        }


def iter_loinc_terms(loinc_table: Path | None, loinc_zh: Path | None) -> Iterable[dict[str, Any]]:
    preferred_by_code: dict[str, str | None] = {}
    aliases_by_code: dict[str, set[str]] = {}
    fallback_by_code: dict[str, str] = {}

    if loinc_table and loinc_table.exists():
        for row in read_table(loinc_table):
            code = pick(row, ("LOINC_NUM", "loinc_num", "loinc", "code"))
            name = pick(row, ("LONG_COMMON_NAME", "long_common_name", "COMPONENT", "component", "name"))
            if code and name:
                fallback_by_code.setdefault(code, name)
                preferred_by_code.setdefault(code, None)
                aliases_by_code.setdefault(code, set()).add(name)

    if loinc_zh and loinc_zh.exists():
        for row in read_table(loinc_zh):
            code = pick(row, ("LOINC_NUM", "loinc_num", "loinc", "code"))
            if not code:
                continue

            long_common_name = pick(row, ("LONG_COMMON_NAME", "long_common_name", "zh_name", "name", "ChineseName", "中文名", "名称"))
            display_name = pick(row, ("LinguisticVariantDisplayName", "DisplayName", "display_name"))
            short_name = pick(row, ("SHORTNAME", "ShortName", "short_name", "短名"))
            component = pick(row, ("COMPONENT", "component"))
            aliases = [long_common_name, display_name, short_name, component]

            # Use a short Chinese name for normalized_name. The full six-axis LOINC
            # definition is useful as an alias/code definition, but is too verbose
            # for AskMed facts and patient-state display.
            for preferred_candidate in (display_name, short_name, component, long_common_name):
                preferred_by_code[code] = choose_better_preferred(preferred_by_code.get(code), preferred_candidate)

            parts = [
                component,
                pick(row, ("PROPERTY", "property")),
                pick(row, ("TIME_ASPCT", "time_aspct")),
                pick(row, ("SYSTEM", "system")),
                pick(row, ("SCALE_TYP", "scale_typ")),
                pick(row, ("METHOD_TYP", "method_typ")),
            ]
            composite = ".".join(part for part in parts if part)
            if composite:
                aliases.append(composite)

            for alias in aliases:
                if alias:
                    aliases_by_code.setdefault(code, set()).add(alias)

    for code, aliases in aliases_by_code.items():
        preferred = preferred_by_code.get(code) or fallback_by_code.get(code) or sorted(aliases)[0]
        yield {
            "terminology": "LOINC",
            "semantic_type": "examination",
            "code": code,
            "preferred_name": preferred,
            "aliases": [preferred, *sorted(aliases - {preferred})],
        }


def init_db(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.execute("drop table if exists terms")
    conn.execute("drop table if exists metadata")
    conn.execute(
        """
        create table terms (
          id integer primary key autoincrement,
          terminology text not null,
          semantic_type text not null,
          code text not null,
          preferred_name text not null,
          alias text not null,
          normalized_key text not null,
          source text
        )
        """
    )
    conn.execute("create index terms_lookup_idx on terms(semantic_type, terminology, normalized_key)")
    conn.execute("create table metadata (key text primary key, value text not null)")
    return conn


def insert_terms(conn: sqlite3.Connection, terms: Iterable[dict[str, Any]], source: str) -> int:
    count = 0
    seen: set[tuple[str, str, str, str]] = set()
    for term in terms:
        for alias in term.get("aliases") or []:
            alias = str(alias or "").strip()
            key = normalize_term_key(alias)
            if not key:
                continue
            identity = (term["terminology"], term["semantic_type"], term["code"], key)
            if identity in seen:
                continue
            seen.add(identity)
            conn.execute(
                """
                insert into terms(terminology, semantic_type, code, preferred_name, alias, normalized_key, source)
                values(?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    term["terminology"],
                    term["semantic_type"],
                    term["code"],
                    term["preferred_name"],
                    alias,
                    key,
                    source,
                ),
            )
            count += 1
    return count


def main() -> None:
    parser = argparse.ArgumentParser(description="Import official terminology files into a local AskMed SQLite index.")
    parser.add_argument("--icd-file", type=Path, default=None)
    parser.add_argument("--loinc-table", type=Path, default=None)
    parser.add_argument("--loinc-zh", type=Path, default=None)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    args.output = resolve_from_root(args.output)
    icd_file = resolve_from_root(args.icd_file) if args.icd_file else None
    loinc_table = resolve_from_root(args.loinc_table) if args.loinc_table else None
    loinc_zh = resolve_from_root(args.loinc_zh) if args.loinc_zh else None
    if not any(path for path in (icd_file, loinc_table, loinc_zh)):
        raise SystemExit("Provide at least one terminology source: --icd-file, --loinc-table, or --loinc-zh.")

    conn = init_db(args.output)
    source_hashes: dict[str, str] = {}
    counts: dict[str, int] = {}
    try:
        if icd_file:
            if not icd_file.exists():
                raise SystemExit(f"ICD file not found: {icd_file}")
            source_hashes[str(icd_file)] = file_sha256(icd_file)
            counts["ICD-10-CN"] = insert_terms(conn, iter_icd_terms(icd_file), str(icd_file))
        if loinc_table or loinc_zh:
            for path in (loinc_table, loinc_zh):
                if path:
                    if not path.exists():
                        raise SystemExit(f"LOINC file not found: {path}")
                    source_hashes[str(path)] = file_sha256(path)
            counts["LOINC"] = insert_terms(conn, iter_loinc_terms(loinc_table, loinc_zh), "LOINC")
        metadata = {
            "created_at": datetime.now(timezone.utc).isoformat(),
            "source_hashes": json.dumps(source_hashes, ensure_ascii=False, sort_keys=True),
            "counts": json.dumps(counts, ensure_ascii=False, sort_keys=True),
        }
        for key, value in metadata.items():
            conn.execute("insert into metadata(key, value) values(?, ?)", (key, value))
        conn.commit()
    finally:
        conn.close()

    print(json.dumps({"output": str(args.output), "counts": counts}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
